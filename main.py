from datetime import datetime
from html import escape
from io import BytesIO
import os
from zoneinfo import ZoneInfo

import bcrypt
import psycopg
from fastapi import Cookie, FastAPI, File, Form, Query, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from itsdangerous import BadSignature, URLSafeSerializer
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

app = FastAPI()

DATABASE_URL = os.getenv("DATABASE_URL")
SECRET_KEY = os.getenv("SECRET_KEY", "cambia-esto")
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD_HASH = os.getenv("ADMIN_PASSWORD_HASH", "")

serializer = URLSafeSerializer(SECRET_KEY, salt="admin-session")


# ---------- Infraestructura ----------
def conectar():
    return psycopg.connect(DATABASE_URL)


def crear_tablas():
    with conectar() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS condominios (
                    id SERIAL PRIMARY KEY,
                    nombre TEXT NOT NULL,
                    slug TEXT UNIQUE NOT NULL,
                    activo BOOLEAN DEFAULT TRUE,
                    creado_en TIMESTAMP DEFAULT NOW()
                )
                """
            )
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS departamentos (
                    id SERIAL PRIMARY KEY,
                    torre TEXT,
                    numero TEXT NOT NULL,
                    UNIQUE(torre, numero)
                )
                """
            )
            cursor.execute("ALTER TABLE departamentos ADD COLUMN IF NOT EXISTS condominio_id INTEGER REFERENCES condominios(id)")
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS residentes (
                    id SERIAL PRIMARY KEY,
                    nombre TEXT NOT NULL,
                    telefono TEXT,
                    email TEXT,
                    tipo TEXT,
                    departamento_id INTEGER REFERENCES departamentos(id)
                )
                """
            )
            cursor.execute("ALTER TABLE residentes ADD COLUMN IF NOT EXISTS condominio_id INTEGER REFERENCES condominios(id)")
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS vehiculos (
                    id SERIAL PRIMARY KEY,
                    patente TEXT NOT NULL,
                    marca TEXT,
                    modelo TEXT,
                    color TEXT,
                    estacionamiento TEXT,
                    departamento_id INTEGER REFERENCES departamentos(id)
                )
                """
            )
            cursor.execute("ALTER TABLE vehiculos ADD COLUMN IF NOT EXISTS estacionamiento TEXT")
            cursor.execute("ALTER TABLE vehiculos ADD COLUMN IF NOT EXISTS condominio_id INTEGER REFERENCES condominios(id)")
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS visitas (
                    id SERIAL PRIMARY KEY,
                    nombre TEXT NOT NULL,
                    rut TEXT,
                    patente TEXT,
                    departamento_id INTEGER REFERENCES departamentos(id),
                    autorizado_por TEXT,
                    observacion TEXT,
                    hora_ingreso TIMESTAMP DEFAULT NOW(),
                    hora_salida TIMESTAMP
                )
                """
            )
            cursor.execute("ALTER TABLE visitas ADD COLUMN IF NOT EXISTS patente TEXT")
            cursor.execute("ALTER TABLE visitas ADD COLUMN IF NOT EXISTS condominio_id INTEGER REFERENCES condominios(id)")
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS encomiendas (
                    id SERIAL PRIMARY KEY,
                    nombre_receptor TEXT NOT NULL,
                    departamento_id INTEGER REFERENCES departamentos(id),
                    descripcion TEXT,
                    recibido_por TEXT,
                    fecha_recepcion TIMESTAMP NOT NULL,
                    fecha_entrega TIMESTAMP,
                    entregado BOOLEAN NOT NULL DEFAULT FALSE,
                    entregado_a TEXT,
                    observacion TEXT
                )
                """
            )
            cursor.execute("ALTER TABLE encomiendas ADD COLUMN IF NOT EXISTS condominio_id INTEGER REFERENCES condominios(id)")
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS usuarios (
                    id SERIAL PRIMARY KEY,
                    username TEXT NOT NULL,
                    password_hash TEXT NOT NULL,
                    rol TEXT NOT NULL,
                    activo BOOLEAN DEFAULT TRUE,
                    creado_en TIMESTAMP DEFAULT NOW()
                )
                """
            )
            cursor.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS condominio_id INTEGER REFERENCES condominios(id)")
            cursor.execute("ALTER TABLE usuarios DROP CONSTRAINT IF EXISTS usuarios_username_key")
            cursor.execute(
                """
                CREATE UNIQUE INDEX IF NOT EXISTS ux_usuarios_condominio_username
                ON usuarios (condominio_id, username)
                """
            )
            cursor.execute(
                """
                CREATE UNIQUE INDEX IF NOT EXISTS ux_departamentos_condominio_torre_numero
                ON departamentos (condominio_id, torre, numero)
                """
            )

            cursor.execute("SELECT id FROM condominios WHERE slug = 'demo'")
            demo = cursor.fetchone()
            if demo:
                demo_id = demo[0]
            else:
                cursor.execute(
                    """
                    INSERT INTO condominios (nombre, slug, activo)
                    VALUES (%s, %s, TRUE)
                    RETURNING id
                    """,
                    ("Condominio Demo", "demo"),
                )
                demo_id = cursor.fetchone()[0]

            cursor.execute("UPDATE departamentos SET condominio_id = %s WHERE condominio_id IS NULL", (demo_id,))
            cursor.execute("UPDATE residentes SET condominio_id = %s WHERE condominio_id IS NULL", (demo_id,))
            cursor.execute("UPDATE vehiculos SET condominio_id = %s WHERE condominio_id IS NULL", (demo_id,))
            cursor.execute("UPDATE visitas SET condominio_id = %s WHERE condominio_id IS NULL", (demo_id,))
            cursor.execute("UPDATE encomiendas SET condominio_id = %s WHERE condominio_id IS NULL", (demo_id,))
            cursor.execute("UPDATE usuarios SET condominio_id = %s WHERE condominio_id IS NULL", (demo_id,))

            cursor.execute("SELECT COUNT(*) FROM usuarios WHERE condominio_id = %s", (demo_id,))
            total_usuarios = cursor.fetchone()[0]
            if total_usuarios == 0 and ADMIN_PASSWORD_HASH:
                cursor.execute(
                    """
                    INSERT INTO usuarios (username, password_hash, rol, activo, condominio_id)
                    VALUES (%s, %s, %s, TRUE, %s)
                    ON CONFLICT (condominio_id, username) DO NOTHING
                    """,
                    (ADMIN_USERNAME, ADMIN_PASSWORD_HASH, "admin", demo_id),
                )
        conn.commit()


# ---------- Auth admin ----------
def crear_token_sesion(username: str, rol: str, condominio_id: int, condominio_nombre: str, condominio_slug: str):
    return serializer.dumps(
        {
            "username": username,
            "rol": rol,
            "condominio_id": condominio_id,
            "condominio_nombre": condominio_nombre,
            "condominio_slug": condominio_slug,
        }
    )


def require_login(token: str | None):
    if not token:
        return None
    try:
        data = serializer.loads(token)
        username = data.get("username")
        rol = data.get("rol")
        condominio_id = data.get("condominio_id")
        condominio_nombre = data.get("condominio_nombre")
        condominio_slug = data.get("condominio_slug")
        if username and rol and condominio_id:
            return {
                "username": username,
                "rol": rol,
                "condominio_id": condominio_id,
                "condominio_nombre": condominio_nombre or "",
                "condominio_slug": condominio_slug or "demo",
            }
    except BadSignature:
        return None
    return None


def verificar_password_admin(password: str):
    if not ADMIN_PASSWORD_HASH:
        return False
    return bcrypt.checkpw(password.encode("utf-8"), ADMIN_PASSWORD_HASH.encode("utf-8"))


def puede_admin(usuario):
    return bool(usuario and usuario.get("rol") == "admin")


def puede_superadmin(usuario):
    return bool(usuario and usuario.get("rol") == "superadmin")


def puede_guardia(usuario):
    return bool(usuario and usuario.get("rol") in {"admin", "guardia"})


def puede_comite(usuario):
    return bool(usuario and usuario.get("rol") in {"admin", "comite"})


def puede_ver_dashboard(usuario):
    return bool(usuario and usuario.get("rol") in {"admin", "guardia", "comite"})


def puede_ver_residentes(usuario):
    return bool(usuario and usuario.get("rol") in {"admin", "guardia", "comite"})


def puede_ver_vehiculos(usuario):
    return bool(usuario and usuario.get("rol") in {"admin", "guardia", "comite"})


def puede_escribir_residentes(usuario):
    return bool(usuario and usuario.get("rol") == "admin")


def puede_escribir_vehiculos(usuario):
    return bool(usuario and usuario.get("rol") == "admin")


def puede_exportar(usuario):
    return bool(usuario and usuario.get("rol") in {"admin", "comite"})


def puede_escribir_visitas(usuario):
    return bool(usuario and usuario.get("rol") in {"admin", "guardia"})


def puede_escribir_encomiendas(usuario):
    return bool(usuario and usuario.get("rol") in {"admin", "guardia"})


def no_permisos_response(usuario):
    if not usuario:
        return RedirectResponse(url="/admin/login", status_code=303)
    contenido = """
    <div class="card">
        <h2>No tienes permisos para esta acción</h2>
        <p class="muted">Tu rol actual no permite ejecutar esta operación.</p>
        <div class="actions"><a class="btn" href="/">Volver al inicio</a></div>
    </div>
    """
    return HTMLResponse(layout("Sin permisos", contenido, usuario))


def condominio_actual_id(usuario):
    return int(usuario.get("condominio_id", 0)) if usuario else 0


# ---------- Helpers UI ----------
def h(value):
    return escape(str(value or ""))


def format_depto(torre, numero):
    if torre and numero:
        return f"{h(torre)}-{h(numero)}"
    return h(torre or numero)


def render_delete_action(es_admin: bool, href: str, confirm_text: str):
    if not es_admin:
        return "<span class='badge warning'>Solo admin</span>"
    return (
        f"<a class='btn red' href='{h(href)}' "
        f"onclick=\"return confirm('{h(confirm_text)}')\">Eliminar</a>"
    )


def ahora_chile() -> datetime:
    return datetime.now(ZoneInfo("America/Santiago")).replace(tzinfo=None)


def badge_estado(texto: str, estilo: str = "neutral"):
    return f"<span class='badge {h(estilo)}'>{h(texto)}</span>"


def encabezados_normalizados(values):
    return [str(v or "").strip().lower() for v in values]


def render_resultado_importacion(titulo: str, volver_url: str, importados: int, omitidos: int, errores: list[str], usuario):
    errores_html = ""
    if errores:
        items = "".join(f"<li>{h(e)}</li>" for e in errores[:80])
        extra = f"<p class='muted'>Mostrando 80 de {len(errores)} errores.</p>" if len(errores) > 80 else ""
        errores_html = f"<h3>Errores</h3><ul>{items}</ul>{extra}"
    contenido = f"""
    <div class="hero"><h1>Importación completada</h1><p>Resultado del proceso de carga masiva.</p></div>
    <div class="card">
        <h2>{h(titulo)}</h2>
        <p><strong>{importados}</strong> registros importados.</p>
        <p><strong>{omitidos}</strong> filas omitidas.</p>
        {errores_html}
        <div class="actions">
            <a class="btn" href="{h(volver_url)}">Volver</a>
        </div>
    </div>
    """
    return HTMLResponse(layout("Importación", contenido, usuario))


def layout(titulo: str, contenido: str, usuario=None):
    usuario_label = "Sesión invitado"
    usuario_badge = "badge dark"
    condominio_label = "Sin condominio"
    admin_link = ""
    if usuario:
        usuario_label = f"{h(usuario.get('username'))} · {h(usuario.get('rol'))}"
        condominio_label = h(usuario.get("condominio_nombre") or "Condominio")
        usuario_badge = "badge info"
        if usuario.get("rol") == "admin":
            admin_link = '<a href="/admin/usuarios">👤 Usuarios</a><a href="/admin/restablecer">🧨 Restablecer datos</a>'
        if usuario.get("rol") == "superadmin":
            admin_link += '<a href="/superadmin">🛡️ Superadmin</a>'
    return f"""
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>{h(titulo)}</title>
        <style>
            :root {{
                --bg:#f1f5f9;
                --surface:#ffffff;
                --text:#0f172a;
                --muted:#64748b;
                --primary:#1d4ed8;
                --primary-dark:#1e3a8a;
                --danger:#dc2626;
                --success:#16a34a;
                --border:#dbe2ea;
                --shadow:0 10px 28px rgba(15,23,42,.08);
            }}
            * {{ box-sizing: border-box; }}
            body {{
                font-family: Inter, ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
                background: var(--bg);
                margin: 0;
                color: var(--text);
            }}
            .app-layout {{
                display: grid;
                grid-template-columns: 260px 1fr;
                min-height: 100vh;
            }}
            .sidebar {{
                background: #0f172a;
                color: #e2e8f0;
                padding: 24px 16px;
                position: sticky;
                top: 0;
                height: 100vh;
            }}
            .logo {{
                display:flex;
                align-items:center;
                gap:10px;
                margin-bottom: 18px;
            }}
            .logo-icon {{
                width: 34px;
                height: 34px;
                border-radius: 10px;
                background: linear-gradient(145deg, #2563eb, #1e40af);
                display:flex;
                align-items:center;
                justify-content:center;
                font-size: 1.05rem;
                box-shadow: 0 8px 16px rgba(37,99,235,.3);
            }}
            .logo-text {{ font-size: 1.1rem; font-weight: 800; color: #fff; letter-spacing:.01em; }}
            .sidebar a {{
                display: block;
                padding: 10px 12px;
                border-radius: 10px;
                color: #cbd5e1;
                text-decoration: none;
                margin-bottom: 8px;
                font-weight: 600;
            }}
            .sidebar a:hover {{ background: #1e293b; color: #fff; }}
            .content-area {{ padding: 20px; }}
            .topbar {{
                background: #fff;
                border: 1px solid var(--border);
                border-radius: 14px;
                padding: 14px 18px;
                margin-bottom: 18px;
                display:flex;
                align-items:center;
                justify-content:space-between;
                box-shadow: var(--shadow);
            }}
            .wrap {{ max-width: 1200px; margin: 0 auto; }}
            .hero {{
                background: linear-gradient(125deg,var(--primary),var(--primary-dark));
                color: white;
                padding: 28px;
                border-radius: 18px;
                margin-bottom: 20px;
                box-shadow: var(--shadow);
            }}
            .hero h1 {{ margin: 0 0 6px; }}
            .hero p {{ margin: 0; opacity: .92; }}
            .card {{
                background: var(--surface);
                border: 1px solid var(--border);
                padding: 22px;
                border-radius: 16px;
                box-shadow: var(--shadow);
                margin-bottom: 18px;
            }}
            .actions {{ display:flex; flex-wrap: wrap; gap: 8px; margin-top: 8px; }}
            form {{ display: grid; grid-template-columns: repeat(auto-fit,minmax(210px,1fr)); gap: 10px; }}
            form .full {{ grid-column: 1/-1; }}
            input, select, textarea {{
                width: 100%;
                padding: 10px 12px;
                border-radius: 10px;
                border: 1px solid #cbd5e1;
                outline: none;
                font-size: .95rem;
            }}
            input:focus, select:focus, textarea:focus {{ border-color: var(--primary); box-shadow: 0 0 0 3px rgba(37,99,235,.15); }}
            button, .btn {{
                display:inline-flex;
                align-items:center;
                justify-content:center;
                background: var(--primary);
                color:#fff;
                border:none;
                border-radius: 10px;
                text-decoration:none;
                padding: 9px 14px;
                font-weight: 600;
                cursor:pointer;
            }}
            .btn.dark {{ background:#0f172a; }}
            .btn.red {{ background: var(--danger); }}
            .btn.green {{ background: var(--success); }}
            .muted {{ color: var(--muted); font-weight: 600; }}
            .table-wrap {{ overflow-x:auto; border: 1px solid var(--border); border-radius: 14px; }}
            table {{ width:100%; border-collapse: separate; border-spacing: 0; min-width: 760px; }}
            th {{ background: #e2e8f0; color: #1e293b; text-align:left; padding: 12px; font-size:.84rem; text-transform: uppercase; letter-spacing: .03em; }}
            td {{ background:#fff; padding: 10px 11px; border-top: 1px solid var(--border); vertical-align: top; }}
            tr:hover td {{ background: #f8fafc; }}
            .badge {{
                display:inline-flex;
                align-items:center;
                padding:4px 10px;
                border-radius: 999px;
                font-size: .76rem;
                font-weight: 700;
            }}
            .badge.success {{ background:#dcfce7; color:#166534; }}
            .badge.warning {{ background:#fef3c7; color:#92400e; }}
            .badge.info {{ background:#dbeafe; color:#1e40af; }}
            .badge.dark {{ background:#e2e8f0; color:#0f172a; }}
            .grid {{ display:grid; grid-template-columns: repeat(auto-fit,minmax(180px,1fr)); gap: 14px; }}
            .stat {{ font-size: 2rem; font-weight: 800; color: var(--primary); }}
            .metric-emoji {{ font-size: 1.3rem; }}
            label {{ font-weight: 600; color:#334155; display:flex; flex-direction:column; gap:6px; }}
            @media (max-width: 980px) {{
                .app-layout {{ grid-template-columns: 1fr; }}
                .sidebar {{ position: static; height: auto; }}
                .content-area {{ padding: 14px; }}
            }}
        </style>
    </head>
    <body>
        <div class="app-layout">
            <aside class="sidebar">
                <div class="logo">
                    <div class="logo-icon">🏢</div>
                    <div class="logo-text">CondoControl<br><small style="font-size:12px;color:#94a3b8;">{condominio_label}</small></div>
                </div>
                <a href="/dashboard-condominio">📊 Dashboard</a>
                <a href="/residentes">👥 Residentes</a>
                <a href="/vehiculos">🚗 Vehículos</a>
                <a href="/visitas">🛂 Visitas</a>
                <a href="/encomiendas">📦 Encomiendas</a>
                {admin_link}
                <a href="/admin/login">🔐 Admin</a>
            </aside>
            <main class="content-area">
                <div class="wrap">
                    <div class="topbar">
                        <strong>{h(titulo)}</strong>
                        <span id="admin-status" class="{usuario_badge}">{usuario_label}</span>
                    </div>
                    {contenido}
                </div>
            </main>
        </div>
    </body>
    </html>
    """


@app.on_event("startup")
def startup_event():
    try:
        crear_tablas()
    except Exception as exc:
        print(f"[startup] No se pudieron crear/verificar tablas: {exc}")


@app.get("/", response_class=HTMLResponse)
def inicio(admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    contenido = """
    <div class="hero">
        <h1>Panel principal</h1>
        <p>Operación diaria del condominio en un solo lugar.</p>
    </div>
    <div class="card">
        <h2>Menú principal</h2>
        <div class="actions">
            <a class="btn" href="/residentes">Residentes</a>
            <a class="btn" href="/vehiculos">Vehículos</a>
            <a class="btn" href="/visitas">Control de visitas</a>
            <a class="btn" href="/encomiendas">Encomiendas</a>
            <a class="btn" href="/dashboard-condominio">Dashboard</a>
            <a class="btn dark" href="/admin/login">Acceso administrador</a>
        </div>
    </div>
    """
    return layout("CondoControl", contenido, usuario)


def render_login_form(condominio_slug: str, condominio_nombre: str):
    contenido = f"""
    <div class="card" style="max-width:460px;margin:auto;">
        <h2>Acceso · {h(condominio_nombre)}</h2>
        <p class="muted">Condominio: <strong>{h(condominio_slug)}</strong></p>
        <form action="/c/{h(condominio_slug)}/login" method="post">
            <label>Usuario<input name="username" placeholder="Usuario" required></label>
            <label>Contraseña<input name="password" type="password" placeholder="Contraseña" required></label>
            <button class="full" type="submit">Entrar</button>
        </form>
        <div class="actions"><a class="btn dark" href="/">Volver</a></div>
    </div>
    """
    return layout("Login", contenido)


@app.get("/admin/login", response_class=HTMLResponse)
def admin_login_form():
    return RedirectResponse(url="/c/demo/login", status_code=303)


@app.get("/c/{slug}/login", response_class=HTMLResponse)
def condominio_login_form(slug: str):
    with conectar() as conn:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, nombre, activo FROM condominios WHERE slug = %s", (slug,))
            condo = cursor.fetchone()
    if not condo or not condo[2]:
        return HTMLResponse("<h3>Condominio no disponible.</h3>", status_code=404)
    return render_login_form(slug, condo[1])


def login_en_condominio(slug: str, username: str, password: str):
    usuario_db = None
    condo = None
    with conectar() as conn:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, nombre, slug, activo FROM condominios WHERE slug = %s", (slug,))
            condo = cursor.fetchone()
            if not condo or not condo[3]:
                return HTMLResponse("<h3>Condominio no disponible.</h3>", status_code=404)
            cursor.execute(
                """
                SELECT username, password_hash, rol, activo
                FROM usuarios
                WHERE username = %s AND condominio_id = %s
                """,
                (username, condo[0]),
            )
            usuario_db = cursor.fetchone()

    login_ok = False
    rol = "admin"
    if usuario_db and usuario_db[3]:
        login_ok = bcrypt.checkpw(password.encode("utf-8"), usuario_db[1].encode("utf-8"))
        rol = usuario_db[2]

    if not login_ok:
        return HTMLResponse(f"<h3>Credenciales incorrectas</h3><a href='/c/{h(slug)}/login'>Volver</a>", status_code=401)

    response = RedirectResponse(url="/", status_code=303)
    response.set_cookie(
        key="admin_session",
        value=crear_token_sesion(
            username=username,
            rol=rol,
            condominio_id=condo[0],
            condominio_nombre=condo[1],
            condominio_slug=condo[2],
        ),
        httponly=True,
        samesite="lax",
        secure=False,
    )
    return response


@app.post("/admin/login")
def admin_login(username: str = Form(...), password: str = Form(...)):
    return login_en_condominio("demo", username, password)


@app.post("/c/{slug}/login")
def condominio_login(slug: str, username: str = Form(...), password: str = Form(...)):
    return login_en_condominio(slug, username, password)


@app.get("/admin/logout")
def admin_logout():
    response = RedirectResponse(url="/", status_code=303)
    response.delete_cookie("admin_session")
    return response


@app.get("/superadmin", response_class=HTMLResponse)
def superadmin_panel(admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    if not puede_superadmin(usuario):
        return no_permisos_response(usuario)
    return RedirectResponse(url="/superadmin/condominios", status_code=303)


@app.get("/superadmin/condominios", response_class=HTMLResponse)
def superadmin_condominios(admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    if not puede_superadmin(usuario):
        return no_permisos_response(usuario)
    with conectar() as conn:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, nombre, slug, activo, creado_en FROM condominios ORDER BY id ASC")
            data = cursor.fetchall()
    filas = "".join(
        f"<tr><td>{c[0]}</td><td>{h(c[1])}</td><td>{h(c[2])}</td><td>{badge_estado('Activo','success') if c[3] else badge_estado('Inactivo','warning')}</td><td>{h(c[4])}</td><td><a class='btn dark' href='/superadmin/condominios/toggle/{c[0]}'>Toggle</a> <a class='btn' href='/superadmin/condominios/{c[0]}/usuarios'>Usuarios</a></td></tr>"
        for c in data
    )
    contenido = f"""
    <div class="hero"><h1>Superadmin</h1><p>Gestión de condominios.</p></div>
    <div class="card">
        <h2>Crear condominio</h2>
        <form action="/superadmin/condominios/crear" method="post">
            <label>Nombre<input name="nombre" required></label>
            <label>Slug<input name="slug" required placeholder="mi-condominio"></label>
            <button class="full" type="submit">Crear</button>
        </form>
    </div>
    <div class="card"><h2>Listado condominios</h2><div class="table-wrap"><table>
    <tr><th>ID</th><th>Nombre</th><th>Slug</th><th>Estado</th><th>Creado</th><th>Acción</th></tr>
    {filas}
    </table></div></div>
    """
    return layout("Superadmin", contenido, usuario)


@app.post("/superadmin/condominios/crear")
def superadmin_condominios_crear(
    admin_session: str | None = Cookie(default=None),
    nombre: str = Form(...),
    slug: str = Form(...),
):
    usuario = require_login(admin_session)
    if not puede_superadmin(usuario):
        return no_permisos_response(usuario)
    with conectar() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                "INSERT INTO condominios (nombre, slug, activo) VALUES (%s, %s, TRUE) ON CONFLICT (slug) DO NOTHING",
                (nombre.strip(), slug.strip().lower()),
            )
        conn.commit()
    return RedirectResponse(url="/superadmin/condominios", status_code=303)


@app.get("/superadmin/condominios/toggle/{condominio_id}")
def superadmin_condominios_toggle(condominio_id: int, admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    if not puede_superadmin(usuario):
        return no_permisos_response(usuario)
    with conectar() as conn:
        with conn.cursor() as cursor:
            cursor.execute("UPDATE condominios SET activo = NOT activo WHERE id = %s", (condominio_id,))
        conn.commit()
    return RedirectResponse(url="/superadmin/condominios", status_code=303)


@app.get("/superadmin/condominios/{condominio_id}/usuarios", response_class=HTMLResponse)
def superadmin_condominio_usuarios(condominio_id: int, admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    if not puede_superadmin(usuario):
        return no_permisos_response(usuario)
    with conectar() as conn:
        with conn.cursor() as cursor:
            cursor.execute("SELECT nombre, slug FROM condominios WHERE id = %s", (condominio_id,))
            condo = cursor.fetchone()
            cursor.execute(
                "SELECT id, username, rol, activo, creado_en FROM usuarios WHERE condominio_id = %s ORDER BY id ASC",
                (condominio_id,),
            )
            data = cursor.fetchall()
    if not condo:
        return HTMLResponse("Condominio no encontrado", status_code=404)
    filas = "".join(
        f"<tr><td>{u[0]}</td><td>{h(u[1])}</td><td>{h(u[2])}</td><td>{'Activo' if u[3] else 'Inactivo'}</td><td>{h(u[4])}</td></tr>"
        for u in data
    )
    contenido = f"""
    <div class="hero"><h1>Usuarios · {h(condo[0])}</h1><p>Slug: {h(condo[1])}</p></div>
    <div class="card"><div class="table-wrap"><table>
    <tr><th>ID</th><th>Username</th><th>Rol</th><th>Activo</th><th>Creado</th></tr>
    {filas}
    </table></div></div>
    <div class="actions"><a class="btn" href="/superadmin/condominios">Volver</a></div>
    """
    return layout("Superadmin usuarios", contenido, usuario)


@app.get("/admin/usuarios", response_class=HTMLResponse)
def admin_usuarios(admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    if not puede_admin(usuario):
        return no_permisos_response(usuario)

    with conectar() as conn:
        with conn.cursor() as cursor:
            condominio_id = condominio_actual_id(usuario)
            cursor.execute(
                """
                SELECT id, username, rol, activo, creado_en
                FROM usuarios
                WHERE condominio_id = %s
                ORDER BY id ASC
                """,
                (condominio_id,),
            )
            data = cursor.fetchall()

    filas = ""
    for u in data:
        activo_badge = badge_estado("Activo", "success") if u[3] else badge_estado("Inactivo", "warning")
        accion_activo = "Desactivar" if u[3] else "Activar"
        filas += f"""
        <tr>
            <td>{u[0]}</td>
            <td>{h(u[1])}</td>
            <td>
                <form action="/admin/usuarios/rol/{u[0]}" method="post" style="display:flex;gap:8px;align-items:center;">
                    <select name="rol" style="min-width:140px;">
                        <option value="admin" {"selected" if u[2] == "admin" else ""}>admin</option>
                        <option value="guardia" {"selected" if u[2] == "guardia" else ""}>guardia</option>
                        <option value="comite" {"selected" if u[2] == "comite" else ""}>comite</option>
                    </select>
                    <button type="submit">Cambiar rol</button>
                </form>
            </td>
            <td>{activo_badge}</td>
            <td>{h(u[4])}</td>
            <td>
                <a class="btn dark" href="/admin/usuarios/toggle/{u[0]}">{accion_activo}</a>
                <a class="btn red" href="/admin/usuarios/eliminar/{u[0]}"
                   onclick="return confirm('¿Eliminar usuario {h(u[1])}?')">Eliminar</a>
            </td>
        </tr>
        """

    contenido = f"""
    <div class="hero"><h1>Usuarios</h1><p>Gestión de cuentas y roles del sistema.</p></div>
    <div class="card">
        <h2>Crear usuario</h2>
        <form action="/admin/usuarios/crear" method="post">
            <label>Username<input name="username" required placeholder="usuario"></label>
            <label>Password<input type="password" name="password" required placeholder="••••••••"></label>
            <label>Rol
                <select name="rol">
                    <option value="guardia">guardia</option>
                    <option value="comite">comite</option>
                    <option value="admin">admin</option>
                </select>
            </label>
            <button class="full" type="submit">Crear usuario</button>
        </form>
    </div>
    <div class="card">
        <h2>Listado de usuarios</h2>
        <div class="table-wrap"><table>
            <tr><th>ID</th><th>Username</th><th>Rol</th><th>Estado</th><th>Creado en</th><th>Acciones</th></tr>
            {filas}
        </table></div>
    </div>
    <div class="actions">
        <a class="btn" href="/">Inicio</a>
        <a class="btn" href="/dashboard-condominio">Dashboard</a>
    </div>
    """
    return layout("Admin usuarios", contenido, usuario)


@app.post("/admin/usuarios/crear")
def admin_usuarios_crear(
    admin_session: str | None = Cookie(default=None),
    username: str = Form(...),
    password: str = Form(...),
    rol: str = Form(...),
):
    usuario = require_login(admin_session)
    if not puede_admin(usuario):
        return no_permisos_response(usuario)
    if rol not in {"admin", "guardia", "comite"}:
        return HTMLResponse("Rol inválido", status_code=400)

    password_hash = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    with conectar() as conn:
        with conn.cursor() as cursor:
            condominio_id = condominio_actual_id(usuario)
            cursor.execute(
                """
                INSERT INTO usuarios (username, password_hash, rol, activo, condominio_id)
                VALUES (%s, %s, %s, TRUE, %s)
                ON CONFLICT (condominio_id, username) DO NOTHING
                """,
                (username.strip(), password_hash, rol, condominio_id),
            )
        conn.commit()
    return RedirectResponse(url="/admin/usuarios", status_code=303)


@app.get("/admin/usuarios/toggle/{user_id}")
def admin_usuarios_toggle(user_id: int, admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    if not puede_admin(usuario):
        return no_permisos_response(usuario)

    with conectar() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                "SELECT username, activo FROM usuarios WHERE id = %s AND condominio_id = %s",
                (user_id, condominio_actual_id(usuario)),
            )
            user = cursor.fetchone()
            if not user:
                return RedirectResponse(url="/admin/usuarios", status_code=303)
            if user[0] == usuario.get("username"):
                return HTMLResponse("No puedes desactivar tu propio usuario.", status_code=400)
            cursor.execute(
                "UPDATE usuarios SET activo = NOT activo WHERE id = %s AND condominio_id = %s",
                (user_id, condominio_actual_id(usuario)),
            )
        conn.commit()
    return RedirectResponse(url="/admin/usuarios", status_code=303)


@app.post("/admin/usuarios/rol/{user_id}")
def admin_usuarios_cambiar_rol(
    user_id: int,
    admin_session: str | None = Cookie(default=None),
    rol: str = Form(...),
):
    usuario = require_login(admin_session)
    if not puede_admin(usuario):
        return no_permisos_response(usuario)
    if rol not in {"admin", "guardia", "comite"}:
        return HTMLResponse("Rol inválido", status_code=400)

    with conectar() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                "UPDATE usuarios SET rol = %s WHERE id = %s AND condominio_id = %s",
                (rol, user_id, condominio_actual_id(usuario)),
            )
        conn.commit()
    return RedirectResponse(url="/admin/usuarios", status_code=303)


@app.get("/admin/usuarios/eliminar/{user_id}")
def admin_usuarios_eliminar(user_id: int, admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    if not puede_admin(usuario):
        return no_permisos_response(usuario)

    with conectar() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                "SELECT username FROM usuarios WHERE id = %s AND condominio_id = %s",
                (user_id, condominio_actual_id(usuario)),
            )
            user = cursor.fetchone()
            if user and user[0] == usuario.get("username"):
                return HTMLResponse("No puedes eliminar tu propio usuario.", status_code=400)
            cursor.execute(
                "DELETE FROM usuarios WHERE id = %s AND condominio_id = %s",
                (user_id, condominio_actual_id(usuario)),
            )
        conn.commit()
    return RedirectResponse(url="/admin/usuarios", status_code=303)


@app.get("/admin/restablecer", response_class=HTMLResponse)
def admin_restablecer_form(admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    if not puede_admin(usuario):
        return no_permisos_response(usuario)

    contenido = """
    <div class="hero"><h1>Restablecer datos de fábrica</h1><p>Herramienta de preparación antes de producción.</p></div>
    <div class="card" style="border:2px solid #dc2626;">
        <h2 style="color:#b91c1c;">Zona peligrosa</h2>
        <p class="muted">
            Esta acción eliminará permanentemente todos los residentes, vehículos, visitas, encomiendas y departamentos,
            incluyendo datos importados desde Excel. No se eliminarán usuarios.
        </p>
        <form action="/admin/restablecer" method="post">
            <label>Escribe exactamente RESTABLECER para confirmar
                <input name="confirmacion" placeholder="RESTABLECER" required>
            </label>
            <button class="full btn red" type="submit">Restablecer datos</button>
        </form>
    </div>
    <div class="actions"><a class="btn" href="/">Volver</a></div>
    """
    return layout("Restablecer datos", contenido, usuario)


@app.post("/admin/restablecer")
def admin_restablecer(
    admin_session: str | None = Cookie(default=None),
    confirmacion: str = Form(...),
):
    usuario = require_login(admin_session)
    if not puede_admin(usuario):
        return no_permisos_response(usuario)

    if confirmacion.strip() != "RESTABLECER":
        return HTMLResponse(
            layout(
                "Restablecer datos",
                """
                <div class="card" style="border:2px solid #dc2626;">
                    <h2 style="color:#b91c1c;">Confirmación inválida</h2>
                    <p>Debes escribir exactamente <strong>RESTABLECER</strong>.</p>
                    <div class="actions"><a class="btn" href="/admin/restablecer">Volver</a></div>
                </div>
                """,
                usuario,
            ),
            status_code=400,
        )

    tablas_operativas = ("residentes", "vehiculos", "visitas", "encomiendas", "departamentos")
    resumen = {tabla: -1 for tabla in tablas_operativas}
    condominio_id = condominio_actual_id(usuario)
    conn = None
    try:
        conn = conectar()
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM visitas WHERE condominio_id = %s", (condominio_id,))
            cursor.execute("DELETE FROM encomiendas WHERE condominio_id = %s", (condominio_id,))
            cursor.execute("DELETE FROM vehiculos WHERE condominio_id = %s", (condominio_id,))
            cursor.execute("DELETE FROM residentes WHERE condominio_id = %s", (condominio_id,))
            cursor.execute("DELETE FROM departamentos WHERE condominio_id = %s", (condominio_id,))

            for tabla in tablas_operativas:
                cursor.execute(f"SELECT COUNT(*) FROM {tabla} WHERE condominio_id = %s", (condominio_id,))
                resumen[tabla] = cursor.fetchone()[0]

            if any(conteo > 0 for conteo in resumen.values()):
                raise RuntimeError("No fue posible restablecer completamente todas las tablas operativas.")
        conn.commit()
    except Exception as exc:
        if conn:
            conn.rollback()
        return HTMLResponse(
            layout(
                "Restablecer datos",
                f"""
                <div class="card" style="border:2px solid #dc2626;">
                    <h2 style="color:#b91c1c;">Error al restablecer datos</h2>
                    <p>{h(exc)}</p>
                    <div class="actions"><a class="btn" href="/admin/restablecer">Volver</a></div>
                </div>
                """,
                usuario,
            ),
            status_code=500,
        )
    finally:
        if conn:
            conn.close()

    contenido = f"""
    <div class="hero"><h1>Restablecimiento completado</h1><p>Se eliminaron los datos operativos del condominio.</p></div>
    <div class="card">
        <h2>Resumen</h2>
        <ul>
            <li>residentes: <strong>{resumen['residentes']}</strong></li>
            <li>vehiculos: <strong>{resumen['vehiculos']}</strong></li>
            <li>visitas: <strong>{resumen['visitas']}</strong></li>
            <li>encomiendas: <strong>{resumen['encomiendas']}</strong></li>
            <li>departamentos: <strong>{resumen['departamentos']}</strong></li>
        </ul>
        <p class="muted">Usuarios, credenciales y roles no fueron eliminados.</p>
        <div class="actions"><a class="btn" href="/">Volver al inicio</a></div>
    </div>
    """
    return HTMLResponse(layout("Restablecimiento completado", contenido, usuario))


def obtener_o_crear_departamento(cursor, condominio_id, torre, numero):
    cursor.execute(
        """
        SELECT id FROM departamentos
        WHERE condominio_id = %s
          AND COALESCE(torre, '') = COALESCE(%s, '')
          AND numero = %s
        """,
        (condominio_id, torre, numero),
    )
    dep = cursor.fetchone()
    if dep:
        return dep[0]

    cursor.execute(
        """
        INSERT INTO departamentos (torre, numero, condominio_id)
        VALUES (%s, %s, %s)
        RETURNING id
        """,
        (torre, numero, condominio_id),
    )
    return cursor.fetchone()[0]


@app.get("/residentes", response_class=HTMLResponse)
def residentes(q: str = Query(default=""), admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    if not puede_ver_residentes(usuario):
        return no_permisos_response(usuario)
    es_admin = puede_escribir_residentes(usuario)
    condominio_id = condominio_actual_id(usuario)
    with conectar() as conn:
        with conn.cursor() as cursor:
            where_parts = ["r.condominio_id = %s"]
            params = [condominio_id]
            if q:
                like = f"%{q}%"
                where_parts.append(
                    """
                    (
                        r.nombre ILIKE %s OR
                        r.telefono ILIKE %s OR
                        r.email ILIKE %s OR
                        r.tipo ILIKE %s OR
                        COALESCE(d.torre, '') ILIKE %s OR
                        d.numero ILIKE %s OR
                        (COALESCE(d.torre, '') || '-' || d.numero) ILIKE %s
                    )
                    """
                )
                params.extend([like, like, like, like, like, like, like])
            where_sql = f"WHERE {' AND '.join(where_parts)}" if where_parts else ""
            cursor.execute(
                """
                SELECT r.id, r.nombre, r.telefono, r.email, r.tipo, d.torre, d.numero
                FROM residentes r
                LEFT JOIN departamentos d ON r.departamento_id = d.id
                """
                + where_sql
                + """
                ORDER BY r.id DESC
                LIMIT 200
                """,
                params,
            )
            data = cursor.fetchall()

    filas = "".join(
        f"""
        <tr>
            <td>{h(r[1])}</td>
            <td>{format_depto(r[5], r[6])}</td>
            <td>{h(r[2])}</td>
            <td>{h(r[3])}</td>
            <td>{h(r[4])}</td>
            <td>{render_delete_action(es_admin, f'/eliminar-residente/{r[0]}', '¿Eliminar residente?')}</td>
        </tr>
        """
        for r in data
    )

    logout = '<a class="btn dark" href="/admin/logout">Cerrar sesión admin</a>' if es_admin else ""

    form_html = """
    <div class="hero"><h1>Residentes</h1><p>Gestión de residentes.</p></div>
    <div class="card">
        <h2>Agregar residente</h2>
        <form action="/guardar-residente" method="post">
            <label>Nombre residente<input name="nombre" placeholder="Nombre residente" required></label>
            <label>Teléfono<input name="telefono" placeholder="Teléfono"></label>
            <label>Email<input name="email" placeholder="Email"></label>
            <label>Tipo residente<select name="tipo">
                <option value="Propietario">Propietario</option>
                <option value="Arrendatario">Arrendatario</option>
                <option value="Residente">Residente</option>
            </select></label>
            <label>Torre / Block<input name="torre" placeholder="Torre / Block"></label>
            <label>Departamento<input name="numero" placeholder="Departamento" required></label>
            <button class="full" type="submit">Guardar residente</button>
        </form>
    </div>
    <div class="card">
        <h2>Importar residentes desde Excel</h2>
        <p class="muted">Columnas requeridas: nombre, telefono, email, tipo, torre, numero</p>
        <form action="/importar/residentes" method="post" enctype="multipart/form-data">
            <label>Archivo .xlsx<input type="file" name="archivo" accept=".xlsx" required></label>
            <button class="full" type="submit">Importar residentes</button>
        </form>
    </div>
    """ if es_admin else """
    <div class="hero"><h1>Residentes</h1><p>Vista de solo lectura para comité.</p></div>
    """

    contenido = f"""
    {form_html}
    <div class="card">
        <h2>Buscar y filtrar</h2>
        <form action="/residentes" method="get">
            <label>Búsqueda
                <input name="q" value="{h(q)}" placeholder="Buscar por nombre, teléfono, email, tipo o depto">
            </label>
            <button type="submit">Aplicar filtros</button>
            <a class="btn dark" href="/residentes">Limpiar</a>
        </form>
    </div>
    <div class="card">
        <h2>Listado</h2>
        <div class="table-wrap"><table>
            <tr><th>Nombre</th><th>Depto</th><th>Teléfono</th><th>Email</th><th>Tipo</th><th>Acción</th></tr>
            {filas}
        </table></div>
    </div>
    <div class="actions">
        <a class="btn" href="/">Inicio</a>
        <a class="btn" href="/dashboard-condominio">Dashboard</a>
        {logout}
    </div>
    """
    return layout("Residentes", contenido, usuario)


@app.post("/guardar-residente")
def guardar_residente(
    admin_session: str | None = Cookie(default=None),
    nombre: str = Form(...),
    telefono: str = Form(""),
    email: str = Form(""),
    tipo: str = Form("Residente"),
    torre: str = Form(""),
    numero: str = Form(...),
):
    usuario = require_login(admin_session)
    if not puede_escribir_residentes(usuario):
        return no_permisos_response(usuario)
    with conectar() as conn:
        with conn.cursor() as cursor:
            dep_id = obtener_o_crear_departamento(cursor, condominio_actual_id(usuario), torre, numero)
            cursor.execute(
                """
                INSERT INTO residentes (nombre, telefono, email, tipo, departamento_id, condominio_id)
                VALUES (%s, %s, %s, %s, %s, %s)
                """,
                (nombre, telefono, email, tipo, dep_id, condominio_actual_id(usuario)),
            )
        conn.commit()
    return RedirectResponse(url="/residentes", status_code=303)


@app.post("/importar/residentes")
async def importar_residentes(admin_session: str | None = Cookie(default=None), archivo: UploadFile = File(...)):
    usuario = require_login(admin_session)
    if not puede_escribir_residentes(usuario):
        return no_permisos_response(usuario)

    if not archivo.filename or not archivo.filename.lower().endswith(".xlsx"):
        return HTMLResponse("Archivo inválido. Debe ser .xlsx", status_code=400)

    importados = 0
    omitidos = 0
    errores: list[str] = []

    try:
        contenido = await archivo.read()
        wb = load_workbook(filename=BytesIO(contenido), data_only=True)
        ws = wb.active
    except Exception as exc:
        return HTMLResponse(f"No se pudo leer el archivo: {h(exc)}", status_code=400)

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return HTMLResponse("El archivo está vacío.", status_code=400)

    headers = encabezados_normalizados(rows[0])
    required = ["nombre", "telefono", "email", "tipo", "torre", "numero"]
    if headers[: len(required)] != required and set(required) - set(headers):
        return HTMLResponse("Encabezados inválidos. Usa: nombre, telefono, email, tipo, torre, numero", status_code=400)

    idx = {hname: headers.index(hname) for hname in required}

    with conectar() as conn:
        with conn.cursor() as cursor:
            for fila_num, row in enumerate(rows[1:], start=2):
                try:
                    row = row or ()
                    vals = [row[idx[k]] if idx[k] < len(row) else None for k in required]
                    nombre, telefono, email, tipo, torre, numero = [(str(v).strip() if v is not None else "") for v in vals]

                    if not any([nombre, telefono, email, tipo, torre, numero]):
                        omitidos += 1
                        continue
                    if not nombre or not numero:
                        omitidos += 1
                        errores.append(f"Fila {fila_num}: nombre y numero son obligatorios.")
                        continue
                    if not tipo:
                        tipo = "Residente"

                    dep_id = obtener_o_crear_departamento(cursor, condominio_actual_id(usuario), torre, numero)
                    cursor.execute(
                        """
                        INSERT INTO residentes (nombre, telefono, email, tipo, departamento_id, condominio_id)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        """,
                        (nombre, telefono, email, tipo, dep_id, condominio_actual_id(usuario)),
                    )
                    conn.commit()
                    importados += 1
                except Exception as exc:
                    conn.rollback()
                    omitidos += 1
                    errores.append(f"Fila {fila_num}: {exc}")

    return render_resultado_importacion("Residentes", "/residentes", importados, omitidos, errores, usuario)


@app.get("/eliminar-residente/{residente_id}")
def eliminar_residente(residente_id: int, admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    if not puede_escribir_residentes(usuario):
        return no_permisos_response(usuario)

    with conectar() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                "DELETE FROM residentes WHERE id = %s AND condominio_id = %s",
                (residente_id, condominio_actual_id(usuario)),
            )
        conn.commit()
    return RedirectResponse(url="/residentes", status_code=303)


@app.get("/vehiculos", response_class=HTMLResponse)
def vehiculos(q: str = Query(default=""), admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    if not puede_ver_vehiculos(usuario):
        return no_permisos_response(usuario)
    es_admin = puede_escribir_vehiculos(usuario)
    condominio_id = condominio_actual_id(usuario)
    with conectar() as conn:
        with conn.cursor() as cursor:
            where_parts = ["v.condominio_id = %s"]
            params = [condominio_id]
            if q:
                like = f"%{q}%"
                where_parts.append(
                    """
                    (
                        v.patente ILIKE %s OR
                        v.marca ILIKE %s OR
                        v.modelo ILIKE %s OR
                        v.color ILIKE %s OR
                        v.estacionamiento ILIKE %s OR
                        COALESCE(d.torre, '') ILIKE %s OR
                        d.numero ILIKE %s OR
                        (COALESCE(d.torre, '') || '-' || d.numero) ILIKE %s
                    )
                    """
                )
                params.extend([like, like, like, like, like, like, like, like])
            where_sql = f"WHERE {' AND '.join(where_parts)}" if where_parts else ""
            cursor.execute(
                """
                SELECT v.id, v.patente, v.marca, v.modelo, v.color, v.estacionamiento, d.torre, d.numero
                FROM vehiculos v
                LEFT JOIN departamentos d ON v.departamento_id = d.id
                """
                + where_sql
                + """
                ORDER BY v.id DESC
                LIMIT 200
                """,
                params,
            )
            data = cursor.fetchall()

    filas = "".join(
        f"""
        <tr>
            <td>{h(v[1])}</td><td>{h(v[2])}</td><td>{h(v[3])}</td><td>{h(v[4])}</td><td>{format_depto(v[6], v[7])}</td><td>{h(v[5])}</td>
            <td>{render_delete_action(es_admin, f'/eliminar-vehiculo/{v[0]}', '¿Eliminar vehículo?')}</td>
        </tr>
        """
        for v in data
    )

    form_html = """
    <div class="hero"><h1>Vehículos</h1><p>Registro de vehículos.</p></div>
    <div class="card">
        <h2>Agregar vehículo</h2>
        <form action="/guardar-vehiculo" method="post">
            <label>Patente<input name="patente" placeholder="Patente" required></label>
            <label>Marca<input name="marca" placeholder="Marca"></label>
            <label>Modelo<input name="modelo" placeholder="Modelo"></label>
            <label>Color<input name="color" placeholder="Color"></label>
            <label>Torre / Block<input name="torre" placeholder="Torre / Block"></label>
            <label>Departamento<input name="numero" placeholder="Departamento" required></label>
            <label>Estacionamiento<input name="estacionamiento" placeholder="N° estacionamiento"></label>
            <button class="full" type="submit">Guardar vehículo</button>
        </form>
    </div>
    <div class="card">
        <h2>Importar vehículos desde Excel</h2>
        <p class="muted">Columnas requeridas: patente, marca, modelo, color, torre, numero, estacionamiento (estacionamiento es opcional)</p>
        <form action="/importar/vehiculos" method="post" enctype="multipart/form-data">
            <label>Archivo .xlsx<input type="file" name="archivo" accept=".xlsx" required></label>
            <button class="full" type="submit">Importar vehículos</button>
        </form>
    </div>
    """ if es_admin else """
    <div class="hero"><h1>Vehículos</h1><p>Vista de solo lectura para comité.</p></div>
    """

    contenido = f"""
    {form_html}
    <div class="card">
        <h2>Buscar y filtrar</h2>
        <form action="/vehiculos" method="get">
            <label>Búsqueda
                <input name="q" value="{h(q)}" placeholder="Buscar por patente, marca, modelo, color, estacionamiento o depto">
            </label>
            <button type="submit">Aplicar filtros</button>
            <a class="btn dark" href="/vehiculos">Limpiar</a>
        </form>
    </div>
    <div class="card">
        <h2>Listado vehículos</h2>
        <div class="table-wrap"><table>
            <tr><th>Patente</th><th>Marca</th><th>Modelo</th><th>Color</th><th>Depto</th><th>Estacionamiento</th><th>Acción</th></tr>
            {filas}
        </table></div>
    </div>
    <div class="actions"><a class="btn" href="/">Inicio</a></div>
    """
    return layout("Vehículos", contenido, usuario)


@app.post("/guardar-vehiculo")
def guardar_vehiculo(
    admin_session: str | None = Cookie(default=None),
    patente: str = Form(...),
    marca: str = Form(""),
    modelo: str = Form(""),
    color: str = Form(""),
    torre: str = Form(""),
    numero: str = Form(...),
    estacionamiento: str = Form(""),
):
    usuario = require_login(admin_session)
    if not puede_escribir_vehiculos(usuario):
        return no_permisos_response(usuario)
    with conectar() as conn:
        with conn.cursor() as cursor:
            dep_id = obtener_o_crear_departamento(cursor, condominio_actual_id(usuario), torre, numero)
            cursor.execute(
                """
                INSERT INTO vehiculos (patente, marca, modelo, color, estacionamiento, departamento_id, condominio_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (patente.upper(), marca, modelo, color, estacionamiento, dep_id, condominio_actual_id(usuario)),
            )
        conn.commit()
    return RedirectResponse(url="/vehiculos", status_code=303)


@app.post("/importar/vehiculos")
async def importar_vehiculos(admin_session: str | None = Cookie(default=None), archivo: UploadFile = File(...)):
    usuario = require_login(admin_session)
    if not puede_escribir_vehiculos(usuario):
        return no_permisos_response(usuario)

    if not archivo.filename or not archivo.filename.lower().endswith(".xlsx"):
        return HTMLResponse("Archivo inválido. Debe ser .xlsx", status_code=400)

    importados = 0
    omitidos = 0
    errores: list[str] = []

    try:
        contenido = await archivo.read()
        wb = load_workbook(filename=BytesIO(contenido), data_only=True)
        ws = wb.active
    except Exception as exc:
        return HTMLResponse(f"No se pudo leer el archivo: {h(exc)}", status_code=400)

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return HTMLResponse("El archivo está vacío.", status_code=400)

    headers = encabezados_normalizados(rows[0])
    required = ["patente", "marca", "modelo", "color", "torre", "numero"]
    optional = ["estacionamiento"]
    if set(required) - set(headers):
        return HTMLResponse("Encabezados inválidos. Usa: patente, marca, modelo, color, torre, numero, estacionamiento", status_code=400)

    idx = {hname: headers.index(hname) for hname in required if hname in headers}
    idx_opt = {hname: headers.index(hname) for hname in optional if hname in headers}

    with conectar() as conn:
        with conn.cursor() as cursor:
            for fila_num, row in enumerate(rows[1:], start=2):
                try:
                    row = row or ()
                    vals = [row[idx[k]] if idx[k] < len(row) else None for k in required]
                    patente, marca, modelo, color, torre, numero = [(str(v).strip() if v is not None else "") for v in vals]
                    estacionamiento = ""
                    if "estacionamiento" in idx_opt:
                        val_est = row[idx_opt["estacionamiento"]] if idx_opt["estacionamiento"] < len(row) else None
                        estacionamiento = str(val_est).strip() if val_est is not None else ""

                    if not any([patente, marca, modelo, color, torre, numero]):
                        omitidos += 1
                        continue
                    if not patente or not numero:
                        omitidos += 1
                        errores.append(f"Fila {fila_num}: patente y numero son obligatorios.")
                        continue

                    dep_id = obtener_o_crear_departamento(cursor, condominio_actual_id(usuario), torre, numero)
                    cursor.execute(
                        """
                        INSERT INTO vehiculos (patente, marca, modelo, color, estacionamiento, departamento_id, condominio_id)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """,
                        (patente.upper(), marca, modelo, color, estacionamiento, dep_id, condominio_actual_id(usuario)),
                    )
                    conn.commit()
                    importados += 1
                except Exception as exc:
                    conn.rollback()
                    omitidos += 1
                    errores.append(f"Fila {fila_num}: {exc}")

    return render_resultado_importacion("Vehículos", "/vehiculos", importados, omitidos, errores, usuario)


@app.get("/eliminar-vehiculo/{vehiculo_id}")
def eliminar_vehiculo(vehiculo_id: int, admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    if not puede_escribir_vehiculos(usuario):
        return no_permisos_response(usuario)

    with conectar() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                "DELETE FROM vehiculos WHERE id = %s AND condominio_id = %s",
                (vehiculo_id, condominio_actual_id(usuario)),
            )
        conn.commit()
    return RedirectResponse(url="/vehiculos", status_code=303)


@app.get("/visitas", response_class=HTMLResponse)
def visitas(q: str = Query(default=""), solo_dentro: int = Query(default=0), admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    if not usuario or usuario.get("rol") not in {"admin", "guardia", "comite"}:
        return no_permisos_response(usuario)
    puede_escribir = puede_escribir_visitas(usuario)
    condominio_id = condominio_actual_id(usuario)
    with conectar() as conn:
        with conn.cursor() as cursor:
            where_parts = ["v.condominio_id = %s"]
            params = [condominio_id]
            if q:
                like = f"%{q}%"
                where_parts.append(
                    """
                    (
                        v.nombre ILIKE %s OR
                        v.rut ILIKE %s OR
                        v.patente ILIKE %s OR
                        COALESCE(d.torre, '') ILIKE %s OR
                        d.numero ILIKE %s OR
                        (COALESCE(d.torre, '') || '-' || d.numero) ILIKE %s
                    )
                    """
                )
                params.extend([like, like, like, like, like, like])
            if solo_dentro:
                where_parts.append("v.hora_salida IS NULL")

            where_sql = f"WHERE {' AND '.join(where_parts)}" if where_parts else ""
            cursor.execute(
                """
                SELECT v.id, v.nombre, v.rut, v.patente, d.torre, d.numero, v.autorizado_por,
                       v.observacion, v.hora_ingreso, v.hora_salida
                FROM visitas v
                LEFT JOIN departamentos d ON v.departamento_id = d.id
                """
                + where_sql
                + """
                ORDER BY v.id DESC
                LIMIT 100
                """,
                params,
            )
            data = cursor.fetchall()

    filas = ""
    for v in data:
        estado = badge_estado("Dentro", "success") if v[9] is None else badge_estado("Salió", "dark")
        salida = (
            f"<a class='btn green' href='/salida-visita/{v[0]}'>Marcar salida</a>"
            if v[9] is None and puede_escribir
            else (h(v[9]) if v[9] is not None else badge_estado("Solo lectura", "warning"))
        )
        filas += f"""
        <tr>
            <td>{h(v[1])}</td>
            <td>{h(v[2])}</td>
            <td>{h(v[3])}</td>
            <td>{format_depto(v[4], v[5])}</td>
            <td>{h(v[6])}</td>
            <td>{h(v[8])}</td>
            <td>{estado}</td>
            <td>{salida}</td>
        </tr>
        """

    checked = "checked" if solo_dentro else ""
    form_html = """
    <div class="hero"><h1>Control de visitas</h1><p>Control de accesos.</p></div>
    <div class="card">
        <h2>Registrar ingreso</h2>
        <form action="/guardar-visita" method="post">
            <label>Nombre visita<input name="nombre" placeholder="Nombre visita" required></label>
            <label>RUT / Documento<input name="rut" placeholder="RUT / Documento"></label>
            <label>Patente (opcional)<input name="patente" placeholder="Patente vehículo (opcional)"></label>
            <label>Torre / Block<input name="torre" placeholder="Torre / Block"></label>
            <label>Departamento que visita<input name="numero" placeholder="Departamento que visita" required></label>
            <label>Autorizado por<input name="autorizado_por" placeholder="Autorizado por"></label>
            <label class="full">Observación<textarea name="observacion" placeholder="Observación"></textarea></label>
            <button class="full" type="submit">Registrar ingreso</button>
        </form>
    </div>
    """ if puede_escribir else """
    <div class="hero"><h1>Control de visitas</h1><p>Vista en modo lectura.</p></div>
    """

    export_link = '<a class="btn" href="/exportar/visitas">Exportar visitas</a>' if puede_exportar(usuario) else ""
    contenido = f"""
    {form_html}
    <div class="card">
        <h2>Buscar y filtrar</h2>
        <form action="/visitas" method="get">
            <label>Búsqueda<input name="q" value="{h(q)}" placeholder="Buscar por nombre, RUT, patente o depto"></label>
            <label style="display:flex;align-items:center;gap:8px;padding:8px 4px;">
                <input type="checkbox" name="solo_dentro" value="1" {checked} style="width:auto;">
                Solo visitas dentro del condominio
            </label>
            <button type="submit">Aplicar filtros</button>
            <a class="btn dark" href="/visitas">Limpiar</a>
        </form>
    </div>
    <div class="card">
        <h2>Últimas visitas</h2>
        <div class="table-wrap"><table>
            <tr><th>Visita</th><th>RUT</th><th>Patente</th><th>Depto</th><th>Autoriza</th><th>Ingreso</th><th>Estado</th><th>Salida</th></tr>
            {filas}
        </table></div>
    </div>
    <div class="actions">
        <a class="btn" href="/">Inicio</a>
        {export_link}
    </div>
    """
    return layout("Visitas", contenido, usuario)


@app.post("/guardar-visita")
def guardar_visita(
    admin_session: str | None = Cookie(default=None),
    nombre: str = Form(...),
    rut: str = Form(""),
    patente: str = Form(""),
    torre: str = Form(""),
    numero: str = Form(...),
    autorizado_por: str = Form(""),
    observacion: str = Form(""),
):
    usuario = require_login(admin_session)
    if not puede_escribir_visitas(usuario):
        return no_permisos_response(usuario)
    hora_ingreso = ahora_chile()
    with conectar() as conn:
        with conn.cursor() as cursor:
            dep_id = obtener_o_crear_departamento(cursor, condominio_actual_id(usuario), torre, numero)
            cursor.execute(
                """
                INSERT INTO visitas (nombre, rut, patente, departamento_id, autorizado_por, observacion, hora_ingreso, condominio_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (nombre, rut, patente.upper(), dep_id, autorizado_por, observacion, hora_ingreso, condominio_actual_id(usuario)),
            )
        conn.commit()
    return RedirectResponse(url="/visitas", status_code=303)


@app.get("/salida-visita/{visita_id}")
def salida_visita(visita_id: int, admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    if not puede_escribir_visitas(usuario):
        return no_permisos_response(usuario)
    hora_salida = ahora_chile()
    with conectar() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                UPDATE visitas
                SET hora_salida = %s
                WHERE id = %s AND hora_salida IS NULL AND condominio_id = %s
                """,
                (hora_salida, visita_id, condominio_actual_id(usuario)),
            )
        conn.commit()
    return RedirectResponse(url="/visitas", status_code=303)


@app.get("/encomiendas", response_class=HTMLResponse)
def encomiendas(q: str = Query(default=""), solo_pendientes: int = Query(default=0), admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    if not usuario or usuario.get("rol") not in {"admin", "guardia", "comite"}:
        return no_permisos_response(usuario)
    puede_escribir = puede_escribir_encomiendas(usuario)
    condominio_id = condominio_actual_id(usuario)
    with conectar() as conn:
        with conn.cursor() as cursor:
            where_parts = ["e.condominio_id = %s"]
            params = [condominio_id]
            if q:
                like = f"%{q}%"
                where_parts.append(
                    """
                    (
                        e.nombre_receptor ILIKE %s OR
                        e.descripcion ILIKE %s OR
                        COALESCE(d.torre, '') ILIKE %s OR
                        d.numero ILIKE %s OR
                        (COALESCE(d.torre, '') || '-' || d.numero) ILIKE %s
                    )
                    """
                )
                params.extend([like, like, like, like, like])
            if solo_pendientes:
                where_parts.append("e.entregado = FALSE")

            where_sql = f"WHERE {' AND '.join(where_parts)}" if where_parts else ""
            cursor.execute(
                """
                SELECT e.id, e.nombre_receptor, d.torre, d.numero, e.descripcion, e.recibido_por,
                       e.fecha_recepcion, e.fecha_entrega, e.entregado, e.entregado_a, e.observacion
                FROM encomiendas e
                LEFT JOIN departamentos d ON e.departamento_id = d.id
                """
                + where_sql
                + """
                ORDER BY e.id DESC
                LIMIT 200
                """,
                params,
            )
            data = cursor.fetchall()

    checked = "checked" if solo_pendientes else ""
    filas = ""
    for e in data:
        estado = badge_estado("Entregada", "info") if e[8] else badge_estado("Pendiente", "warning")
        entrega = f"{h(e[7])} - {h(e[9])}" if e[8] else (
            f"""
            <form action="/entregar-encomienda/{e[0]}" method="get" style="display:flex;gap:6px;align-items:center;">
                <input name="entregado_a" placeholder="Entregado a" style="min-width:140px;">
                <button class="btn green" type="submit">Marcar entrega</button>
            </form>
            """
            if puede_escribir
            else badge_estado("Solo lectura", "warning")
        )
        filas += f"""
        <tr>
            <td>{h(e[1])}</td>
            <td>{format_depto(e[2], e[3])}</td>
            <td>{h(e[4])}</td>
            <td>{h(e[5])}</td>
            <td>{h(e[6])}</td>
            <td>{estado}</td>
            <td>{entrega}</td>
            <td>{h(e[10])}</td>
        </tr>
        """

    form_html = """
    <div class="hero"><h1>Encomiendas</h1><p>Gestión de paquetes.</p></div>
    <div class="card">
        <h2>Registrar encomienda</h2>
        <form action="/guardar-encomienda" method="post">
            <label>Nombre receptor<input name="nombre_receptor" placeholder="Nombre receptor" required></label>
            <label>Torre / Block<input name="torre" placeholder="Torre / Block"></label>
            <label>Departamento<input name="numero" placeholder="Departamento" required></label>
            <label>Descripción<input name="descripcion" placeholder="Descripción"></label>
            <label>Recibido por<input name="recibido_por" placeholder="Recibido por (conserje)"></label>
            <label class="full">Observación<textarea name="observacion" placeholder="Observación"></textarea></label>
            <button class="full" type="submit">Guardar encomienda</button>
        </form>
    </div>
    """ if puede_escribir else """
    <div class="hero"><h1>Control de encomiendas</h1><p>Vista en modo lectura.</p></div>
    """

    export_link = '<a class="btn" href="/exportar/encomiendas">Exportar encomiendas</a>' if puede_exportar(usuario) else ""
    contenido = f"""
    {form_html}
    <div class="card">
        <h2>Buscar y filtrar</h2>
        <form action="/encomiendas" method="get">
            <label>Búsqueda<input name="q" value="{h(q)}" placeholder="Buscar por receptor, depto o descripción"></label>
            <label style="display:flex;align-items:center;gap:8px;padding:8px 4px;">
                <input type="checkbox" name="solo_pendientes" value="1" {checked} style="width:auto;">
                Solo pendientes por entregar
            </label>
            <button type="submit">Aplicar filtros</button>
            <a class="btn dark" href="/encomiendas">Limpiar</a>
        </form>
    </div>
    <div class="card">
        <h2>Listado de encomiendas</h2>
        <div class="table-wrap"><table>
            <tr><th>Receptor</th><th>Depto</th><th>Descripción</th><th>Recibido por</th><th>Recepción</th><th>Estado</th><th>Entrega</th><th>Observación</th></tr>
            {filas}
        </table></div>
    </div>
    <div class="actions">
        <a class="btn" href="/">Inicio</a>
        <a class="btn" href="/dashboard-condominio">Dashboard</a>
        {export_link}
    </div>
    """
    return layout("Encomiendas", contenido, usuario)


@app.post("/guardar-encomienda")
def guardar_encomienda(
    admin_session: str | None = Cookie(default=None),
    nombre_receptor: str = Form(...),
    torre: str = Form(""),
    numero: str = Form(...),
    descripcion: str = Form(""),
    recibido_por: str = Form(""),
    observacion: str = Form(""),
):
    usuario = require_login(admin_session)
    if not puede_escribir_encomiendas(usuario):
        return no_permisos_response(usuario)
    fecha_recepcion = ahora_chile()
    with conectar() as conn:
        with conn.cursor() as cursor:
            dep_id = obtener_o_crear_departamento(cursor, condominio_actual_id(usuario), torre, numero)
            cursor.execute(
                """
                INSERT INTO encomiendas (
                    nombre_receptor, departamento_id, descripcion, recibido_por,
                    fecha_recepcion, observacion, condominio_id
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (nombre_receptor, dep_id, descripcion, recibido_por, fecha_recepcion, observacion, condominio_actual_id(usuario)),
            )
        conn.commit()
    return RedirectResponse(url="/encomiendas", status_code=303)


@app.get("/entregar-encomienda/{encomienda_id}")
def entregar_encomienda(encomienda_id: int, entregado_a: str = Query(default=""), admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    if not puede_escribir_encomiendas(usuario):
        return no_permisos_response(usuario)
    fecha_entrega = ahora_chile()
    entregado_a_value = entregado_a.strip() or "Recibido por residente"
    with conectar() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                UPDATE encomiendas
                SET entregado = TRUE, fecha_entrega = %s, entregado_a = %s
                WHERE id = %s AND entregado = FALSE AND condominio_id = %s
                """,
                (fecha_entrega, entregado_a_value, encomienda_id, condominio_actual_id(usuario)),
            )
        conn.commit()
    return RedirectResponse(url="/encomiendas", status_code=303)


@app.get("/dashboard-condominio", response_class=HTMLResponse)
def dashboard_condominio(admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    if not puede_ver_dashboard(usuario):
        return no_permisos_response(usuario)
    condominio_id = condominio_actual_id(usuario)
    with conectar() as conn:
        with conn.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM residentes WHERE condominio_id = %s", (condominio_id,))
            total_residentes = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM vehiculos WHERE condominio_id = %s", (condominio_id,))
            total_vehiculos = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM visitas WHERE condominio_id = %s AND DATE(hora_ingreso) = CURRENT_DATE", (condominio_id,))
            visitas_hoy = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM visitas WHERE condominio_id = %s AND hora_salida IS NULL", (condominio_id,))
            visitas_dentro = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM encomiendas WHERE condominio_id = %s AND entregado = FALSE", (condominio_id,))
            encomiendas_pendientes = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM encomiendas WHERE condominio_id = %s AND DATE(fecha_recepcion) = CURRENT_DATE", (condominio_id,))
            encomiendas_recibidas_hoy = cursor.fetchone()[0]

            cursor.execute(
                "SELECT COUNT(*) FROM encomiendas WHERE condominio_id = %s AND entregado = TRUE AND DATE(fecha_entrega) = CURRENT_DATE",
                (condominio_id,),
            )
            encomiendas_entregadas_hoy = cursor.fetchone()[0]

            cursor.execute(
                """
                SELECT d.torre, d.numero, COUNT(v.id) AS total
                FROM visitas v
                LEFT JOIN departamentos d ON v.departamento_id = d.id
                WHERE v.condominio_id = %s
                GROUP BY d.torre, d.numero
                ORDER BY total DESC
                LIMIT 5
                """,
                (condominio_id,),
            )
            top_deptos = cursor.fetchall()

    top_html = "".join(f"<li>{format_depto(d[0], d[1])}: {d[2]} visitas</li>" for d in top_deptos) or "<li>No hay datos</li>"
    ahora = ahora_chile().strftime("%Y-%m-%d %H:%M")

    contenido = f"""
    <div class="hero"><h1>Dashboard</h1><p>Resumen general del condominio.</p></div>
    <div class="grid">
        <div class="card"><h3><span class="metric-emoji">👥</span> Residentes</h3><div class="stat">{total_residentes}</div></div>
        <div class="card"><h3><span class="metric-emoji">🚗</span> Vehículos</h3><div class="stat">{total_vehiculos}</div></div>
        <div class="card"><h3><span class="metric-emoji">🛂</span> Visitas hoy</h3><div class="stat">{visitas_hoy}</div></div>
        <div class="card"><h3><span class="metric-emoji">🏠</span> Visitas dentro</h3><div class="stat">{visitas_dentro}</div></div>
        <div class="card"><h3><span class="metric-emoji">📦</span> Encomiendas pendientes</h3><div class="stat">{encomiendas_pendientes}</div></div>
        <div class="card"><h3><span class="metric-emoji">📥</span> Recibidas hoy</h3><div class="stat">{encomiendas_recibidas_hoy}</div></div>
        <div class="card"><h3><span class="metric-emoji">✅</span> Entregadas hoy</h3><div class="stat">{encomiendas_entregadas_hoy}</div></div>
    </div>
    <div class="card">
        <h2>Resumen operativo</h2>
        <p class="muted">Visitas activas: {visitas_dentro} · Encomiendas por entregar: {encomiendas_pendientes}.</p>
        <p class="muted">Actualizado: {h(ahora)} (hora local Chile).</p>
    </div>
    <div class="card"><h2>Top departamentos con más visitas</h2><ul>{top_html}</ul><p class="muted">Actualizado: {h(ahora)}</p></div>
    <div class="actions">
        <a class="btn" href="/">Inicio</a>
        <a class="btn" href="/visitas">Control visitas</a>
        <a class="btn" href="/encomiendas">Encomiendas</a>
        <a class="btn" href="/exportar/visitas">Exportar visitas</a>
    </div>
    """
    return layout("Dashboard Condominio", contenido, usuario)


@app.get("/exportar/visitas")
def exportar_visitas(admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    if not puede_exportar(usuario):
        return no_permisos_response(usuario)
    with conectar() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT v.id, v.nombre, v.rut, v.patente, d.torre, d.numero, v.autorizado_por,
                       v.observacion, v.hora_ingreso, v.hora_salida
                FROM visitas v
                LEFT JOIN departamentos d ON v.departamento_id = d.id
                WHERE v.condominio_id = %s
                ORDER BY v.id DESC
                """,
                (condominio_actual_id(usuario),),
            )
            visitas = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Visitas"

    headers = [
        "ID",
        "Nombre visita",
        "RUT",
        "Patente",
        "Torre",
        "Departamento",
        "Autorizado por",
        "Observación",
        "Hora ingreso",
        "Hora salida",
    ]
    ws.append(headers)

    fill = PatternFill(fill_type="solid", fgColor="2563EB")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center")

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align

    for visita in visitas:
        ws.append(list(visita))

    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[col].width = 22

    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)

    return StreamingResponse(
        archivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=visitas_condominio.xlsx"},
    )


@app.get("/exportar/encomiendas")
def exportar_encomiendas(admin_session: str | None = Cookie(default=None)):
    usuario = require_login(admin_session)
    if not puede_exportar(usuario):
        return no_permisos_response(usuario)
    with conectar() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT e.id, e.nombre_receptor, d.torre, d.numero, e.descripcion, e.recibido_por,
                       e.fecha_recepcion, e.entregado, e.fecha_entrega, e.entregado_a, e.observacion
                FROM encomiendas e
                LEFT JOIN departamentos d ON e.departamento_id = d.id
                WHERE e.condominio_id = %s
                ORDER BY e.id DESC
                """,
                (condominio_actual_id(usuario),),
            )
            data = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Encomiendas"

    headers = [
        "ID",
        "Nombre receptor",
        "Torre",
        "Departamento",
        "Descripción",
        "Recibido por",
        "Fecha recepción",
        "Entregado",
        "Fecha entrega",
        "Entregado a",
        "Observación",
    ]
    ws.append(headers)

    fill = PatternFill(fill_type="solid", fgColor="2563EB")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align

    for row in data:
        ws.append(list(row))

    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
        ws.column_dimensions[col].width = 22

    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)

    return StreamingResponse(
        archivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=encomiendas_condominio.xlsx"},
    )


@app.get("/health")
def health():
    return {"ok": True}
